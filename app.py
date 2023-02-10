import csv
import openpyxl
import datetime

user_map = {}
curriculum_map = {}
unique_training_titles = set()
unique_user_id = set()
training_titles = {}

# Read User groups and User IDs from Table 3
with open("Table 3.csv", "r") as file:
    reader = csv.reader(file, delimiter=';')
    next(reader)  # Skip header row
    for row in reader:
        user_id, user_group = row
        if user_id in user_map:
            user_map[user_id].append(user_group)
        else:
            user_map[user_id] = [user_group]

# Read Curriculum and User groups from Table 2
with open("Table 2.csv", "r") as file:
    reader = csv.reader(file, delimiter=';')
    next(reader)  # Skip header row
    for row in reader:
        curriculum, user_group = row
        if user_group in curriculum_map:
            curriculum_map[user_group].append(curriculum)
        else:
            curriculum_map[user_group] = [curriculum]

# Read Curriculum, Training Title and Initial Due from Table 1
with open("Table 1.csv", "r") as file:
    reader = csv.reader(file, delimiter=';')
    next(reader)  # Skip header row
    for row in reader:
        curriculum, title, due_date = row
        if curriculum in curriculum_map:
            curriculum_map[curriculum].append((title, due_date))
        else:
            curriculum_map[curriculum] = [(title, due_date)]

# Get a set of unique User IDs.
with open("Table 3.csv", "r") as file:
    reader = csv.reader(file, delimiter=';')
    next(reader)  # Skip header row
    for row in reader:
        user_id, _ = row
        unique_user_id.add(user_id)

# Get a set of unique Training Titles.
with open("Table 1.csv", "r") as file:
    reader = csv.reader(file, delimiter=';')
    next(reader)  # Skip header row
    for row in reader:
        _, title, _ = row
        unique_training_titles.add(title)

# Store each user's Curriculum and Training Titles
user_training = {}
for user_id, user_groups in user_map.items():
    training = []
    for user_group in user_groups:
        if user_group in curriculum_map:
            curriculums = curriculum_map[user_group]
            for curriculum in curriculums:
                if curriculum in curriculum_map:
                    titles = curriculum_map[curriculum]
                    training.extend(titles)
    user_training[user_id] = training

# Create an Excel workbook and add a worksheet

template = "Training Matrix.xlsx"
workbook = openpyxl.load_workbook(template)

# Get the current date and time
now = datetime.datetime.now()
current_time = now.strftime("%Y-%m-%d %H.%M.%S")

# Create a new file with the current date and time in the filename
new_file = f"{template.split('.')[0]} ({current_time}).xlsx"

worksheet = workbook.active
worksheet.title = "Training Matrix"

# Write the headers to the worksheet
worksheet['A1'] = "User ID"

row = 2
column = 2

# Write the titles in the Training Matrix
for title in sorted(unique_training_titles):
    worksheet.cell(1, column, title)
    column += 1

#Populate User IDs in Column A
for user_id in sorted(unique_user_id):
    worksheet.cell(row, 1, user_id)
    row += 1

# Remove duplicates Training Titles.
for user_id, training in sorted(user_training.items()):
    for t in training:
        if len(t) == 2:
            title, due_date = t
            if title not in training_titles:
                training_titles[title] = due_date
            elif due_date < training_titles[title]:
                training_titles[title] = due_date

def convert_to_int(string_value):
    try:
        return int(string_value)
    except ValueError:
        return string_value

# Loop through all columns & rows and put a zero in them.
for col in range(2, (len(unique_training_titles) + 2)):
    for row in range (2, (len(unique_user_id) + 2)):
        worksheet.cell(row, col, 0)
    
row = 2
# Loop through columns B to BN
for col in range(2, (len(unique_training_titles) + 2)):
    value = worksheet.cell(row=1, column=col).value

    row = 2
    for user_id, training in sorted(user_training.items()):
        for title in training:
            due_date = convert_to_int(title[1])

            if title[0] == value: 
                worksheet.cell(row, col, due_date)
        row += 1

# Save the workbook
workbook.save(new_file)
